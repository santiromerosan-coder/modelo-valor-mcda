"""
Compact summary of each sheet — focused on structure, headers, and formulas.
Produces a human-readable text overview for planning the web app.
"""
import openpyxl
import warnings
warnings.filterwarnings('ignore')

SRC = '/home/z/my-project/upload/MV_naMD_V5_VF_1072026 (1).xlsm'
OUT = '/home/z/my-project/data/sheet_overview.txt'

wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

lines = []

def add(s=''):
    lines.append(s)

for sheet_name in wb_f.sheetnames:
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]
    add('=' * 100)
    add(f'SHEET: "{sheet_name}"  |  state={ws_f.sheet_state}  |  dims={ws_f.dimensions}  |  rows={ws_f.max_row}  cols={ws_f.max_column}')
    add(f'Merged ranges: {len(list(ws_f.merged_cells.ranges))}')
    add('=' * 100)
    
    # Build a matrix of cells
    from openpyxl.utils import get_column_letter
    max_r = min(ws_f.max_row, 200)
    max_c = min(ws_f.max_column, 80)
    
    # Header row with column letters
    if max_c > 0:
        header = '       ' + ' | '.join(f'{get_column_letter(c):>8s}' for c in range(1, max_c+1))
        add(header)
        add('-' * len(header))
    
    for r in range(1, max_r+1):
        row_cells = []
        for c in range(1, max_c+1):
            cell_f = ws_f.cell(row=r, column=c)
            cell_v = ws_v.cell(row=r, column=c)
            v_f = cell_f.value
            v_v = cell_v.value
            if v_f is None and v_v is None:
                row_cells.append('        ')
            else:
                # Use formula if available, else value
                if isinstance(v_f, str) and v_f.startswith('='):
                    display = v_f[:8]
                else:
                    val = v_v if v_v is not None else v_f
                    s = str(val)
                    if len(s) > 8:
                        s = s[:7] + '…'
                    display = s
                row_cells.append(f'{display:>8s}')
        add(f'R{r:>4d}: ' + ' | '.join(row_cells))
    
    # Show all formulas in this sheet
    formulas = []
    for row in ws_f.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append((cell.coordinate, cell.value))
    if formulas:
        add(f'\n--- Formulas ({len(formulas)} total) ---')
        for coord, f in formulas[:80]:
            add(f'  {coord}: {f[:120]}')
        if len(formulas) > 80:
            add(f'  ... and {len(formulas)-80} more')
    
    # Charts
    if hasattr(ws_f, '_charts') and ws_f._charts:
        add(f'\n--- Charts: {len(ws_f._charts)} ---')
        for ch in ws_f._charts:
            try:
                title = ch.title
                add(f'  Chart type: {type(ch).__name__}, title: {title}')
            except:
                add(f'  Chart type: {type(ch).__name__}')
    
    # Conditional formatting
    cf = ws_f.conditional_formatting
    cf_rules = list(cf._cf_rules.items()) if hasattr(cf, '_cf_rules') else []
    if cf_rules:
        add(f'\n--- Conditional formatting: {len(cf_rules)} ranges ---')
    
    # Data validations
    dvs = ws_f.data_validations.dataValidation if ws_f.data_validations else []
    if dvs:
        add(f'\n--- Data validations: {len(dvs)} ---')
        for dv in dvs:
            add(f'  cells={dv.sqref} type={dv.type} formula1={dv.formula1}')
    
    add('')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'Saved overview to {OUT}')
print(f'Total lines: {len(lines)}')
