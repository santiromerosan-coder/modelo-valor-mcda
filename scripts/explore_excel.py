"""
Deep exploration of the MV_naMD Excel file.
Extracts: formulas, values, merged cells, charts, styling, defined names, conditional formats.
Saves structured JSON per sheet under /home/z/my-project/data/excel_dump/.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import warnings
warnings.filterwarnings('ignore')

SRC = '/home/z/my-project/upload/MV_naMD_V5_VF_1072026 (1).xlsm'
OUT_DIR = '/home/z/my-project/data/excel_dump'
os.makedirs(OUT_DIR, exist_ok=True)

# Load twice: one with formulas, one with cached values
wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

summary = {
    'sheets': [],
    'defined_names': [],
}

# Defined names
try:
    for name, defn in wb_f.defined_names.items():
        summary['defined_names'].append({
            'name': name,
            'value': str(defn.value) if defn.value else None,
        })
except Exception as e:
    summary['defined_names_error'] = str(e)

for sheet_name in wb_f.sheetnames:
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    sheet_info = {
        'name': sheet_name,
        'state': ws_f.sheet_state,
        'dimensions': ws_f.dimensions,
        'max_row': ws_f.max_row,
        'max_col': ws_f.max_column,
        'merged_ranges': [str(r) for r in ws_f.merged_cells.ranges],
        'cells': [],
    }

    # Iterate cells
    for row in ws_f.iter_rows():
        for cell_f in row:
            v_f = cell_f.value
            v_v = ws_v[cell_f.coordinate].value
            if v_f is None and v_v is None:
                continue

            cell_data = {
                'ref': cell_f.coordinate,
                'formula': v_f if (isinstance(v_f, str) and str(v_f).startswith('=')) else None,
                'raw': v_f if not (isinstance(v_f, str) and str(v_f).startswith('=')) else None,
                'value': v_v,
            }
            # Only keep non-empty cells
            if any(x is not None for x in [cell_data['formula'], cell_data['raw'], cell_data['value']]):
                # Style info
                if cell_f.has_style:
                    fill = cell_f.fill
                    font = cell_f.font
                    cell_data['style'] = {
                        'fill': fill.fgColor.rgb if fill and fill.fgColor and fill.patternType else None,
                        'font_bold': font.b if font else False,
                        'font_color': font.color.rgb if font and font.color else None,
                        'font_size': font.size if font else None,
                        'number_format': cell_f.number_format,
                    }
                sheet_info['cells'].append(cell_data)

    summary['sheets'].append({
        'name': sheet_name,
        'state': sheet_info['state'],
        'dimensions': sheet_info['dimensions'],
        'max_row': sheet_info['max_row'],
        'max_col': sheet_info['max_col'],
        'merged_count': len(sheet_info['merged_ranges']),
        'cell_count': len(sheet_info['cells']),
    })

    # Save individual sheet
    out_path = os.path.join(OUT_DIR, f'{sheet_name.strip().replace(" ", "_").replace("/", "_")}.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(sheet_info, f, ensure_ascii=False, indent=2, default=str)
    print(f'Saved: {out_path}  ({len(sheet_info["cells"])} cells)')

# Save summary
with open(os.path.join(OUT_DIR, '_summary.json'), 'w', encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=2, default=str)

print('\n=== SUMMARY ===')
for s in summary['sheets']:
    print(f'  {s["name"]:35s} | state={s["state"]:10s} | cells={s["cell_count"]:5d} | merged={s["merged_count"]}')
print(f'\nDefined names: {len(summary["defined_names"])}')
for dn in summary['defined_names'][:30]:
    print(f'  {dn["name"]:40s} = {dn["value"]}')
