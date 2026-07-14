"""
Focused inspection of 'Resumen' and 'Referencias' sheets with full values.
Also dumps 'Inputs_nAMD', 'Input_costo2' and the MCDA dashboard with cached values.
"""
import openpyxl
import warnings
import json
warnings.filterwarnings('ignore')

SRC = '/home/z/my-project/upload/MV_naMD_V5_VF_1072026 (1).xlsm'
wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

def dump_sheet(name, max_row=None, max_col=None):
    ws_f = wb_f[name]
    ws_v = wb_v[name]
    mr = max_row or ws_f.max_row
    mc = max_col or ws_f.max_column
    print(f'\n{"="*100}')
    print(f'SHEET: "{name}" rows={mr} cols={mc}')
    print(f'{"="*100}')
    from openpyxl.utils import get_column_letter
    for r in range(1, mr+1):
        row_vals = []
        for c in range(1, mc+1):
            cell_f = ws_f.cell(row=r, column=c)
            cell_v = ws_v.cell(row=r, column=c)
            v_f = cell_f.value
            v_v = cell_v.value
            if v_f is None and v_v is None:
                continue
            # Use formula if formula else value
            if isinstance(v_f, str) and v_f.startswith('='):
                display = f'FORMULA'
                val = v_f
            else:
                val = v_v if v_v is not None else v_f
                display = str(val)[:60]
            row_vals.append(f'{get_column_letter(c)}{r}={display}')
        if row_vals:
            print(f'R{r}: ' + ' | '.join(row_vals))

# Resumen with cached values
print('\n\n############ RESUMEN SHEET ############')
dump_sheet('Resumen', max_row=37, max_col=67)

# Referencias
print('\n\n############ REFERENCIAS SHEET ############')
dump_sheet('Referencias', max_row=70, max_col=3)

# MCDA with cached values
print('\n\n############ MCDA ADVANCED DASHBOARD ############')
dump_sheet('MCDA Advanced Dashboard', max_row=30, max_col=18)

# Inputs_nAMD - just first 2 molecules to see the pattern
print('\n\n############ INPUTS_nAMD (first 75 rows) ############')
dump_sheet('Inputs_nAMD ', max_row=75, max_col=8)
