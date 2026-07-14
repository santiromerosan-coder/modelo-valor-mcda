"""
Extrae todos los datos del NUEVO Excel para comparar con lo que tenemos.
"""
import openpyxl
import json
import warnings
warnings.filterwarnings('ignore')

SRC = '/home/z/my-project/upload/Modelo de valor degeneración macular asociada a la edad.xlsm'
OUT = '/home/z/my-project/data/new_seed_data.json'

wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

def cell_v(sheet, coord):
    return wb_v[sheet][coord].value

# ============================================================
# 1. Molecules (from Modelo de valor)
# ============================================================
molecules = []
mv_sheet = 'Modelo de valor'
for i, r in enumerate(range(13, 21), start=0):
    name = cell_v(mv_sheet, f'N{r}')
    col_range = cell_v(mv_sheet, f'Q{r}')
    selected = cell_v(mv_sheet, f'P{r}')
    molecules.append({
        'id': f'mol_{i}',
        'index': i,
        'name': str(name).strip() if name else '',
        'columnRange': str(col_range).strip() if col_range else '',
        'selected': bool(selected),
    })

discount = cell_v(mv_sheet, 'O24') or 0.0

# ============================================================
# 2. Definitions (from Definiciones sheet)
# ============================================================
definitions = []
def_sheet = 'Definiciones '
for r in range(5, 15):
    domain = cell_v(def_sheet, f'A{r}')
    criteria = cell_v(def_sheet, f'B{r}')
    operational = cell_v(def_sheet, f'C{r}')
    possible_values = cell_v(def_sheet, f'D{r}')
    if criteria:
        definitions.append({
            'domain': str(domain).strip() if domain else '',
            'criteria': str(criteria).strip() if criteria else '',
            'operationalDefinition': str(operational).strip() if operational else '',
            'possibleValues': str(possible_values).strip() if possible_values else '',
        })

# ============================================================
# 3. Parameters (from Parametros sheet)
# ============================================================
parameters = []
param_sheet = 'Parametros '
current_domain = None
current_criteria = None
for r in range(6, 32):
    domain = cell_v(param_sheet, f'A{r}')
    criteria = cell_v(param_sheet, f'B{r}')
    level = cell_v(param_sheet, f'C{r}')
    value = cell_v(param_sheet, f'D{r}')
    if domain:
        current_domain = str(domain).strip()
    if criteria:
        current_criteria = str(criteria).strip()
    if level is not None and value is not None:
        parameters.append({
            'domain': current_domain,
            'criteria': current_criteria,
            'level': str(level).strip() if level else '',
            'value': int(value) if value is not None else 0,
        })

# ============================================================
# 4. Inputs_EMD (outcomes per molecule) — NUEVA HOJA
# ============================================================
inputs_sheet = 'Inputs_EMD '
# Find molecule blocks
molecule_block_starts = []
ws = wb_v[inputs_sheet]
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=3).value  # column C
    if v and isinstance(v, str):
        v_clean = v.strip()
        if any(v_clean == m['name'] for m in molecules):
            molecule_block_starts.append((r, v_clean))

molecule_inputs = {}
for start_row, mname in molecule_block_starts:
    # Find "Desenlace de interés" header
    header_row = None
    for r in range(start_row + 1, start_row + 9):
        v = cell_v(inputs_sheet, f'C{r}')
        if v and 'Desenlace' in str(v):
            header_row = r
            break
    if header_row is None:
        header_row = start_row + 4
    outcomes = []
    for i in range(10):
        r = header_row + 1 + i
        outcome_name = cell_v(inputs_sheet, f'C{r}')
        result = cell_v(inputs_sheet, f'D{r}')
        value_formula = cell_v(inputs_sheet, f'E{r}')
        followup = cell_v(inputs_sheet, f'F{r}')
        reference = cell_v(inputs_sheet, f'G{r}')
        comments = cell_v(inputs_sheet, f'H{r}')
        if outcome_name:
            outcomes.append({
                'outcome': str(outcome_name).strip(),
                'result': str(result).strip() if result is not None else '',
                'value': value_formula if isinstance(value_formula, (int, float)) else None,
                'followup': str(followup).strip() if followup else '',
                'reference': str(reference).strip() if reference else '',
                'comments': str(comments).strip() if comments else '',
            })
    molecule_inputs[mname] = outcomes

# ============================================================
# 5. Cost inputs (from Input_costo2)
# ============================================================
cost_sheet = 'Input_costo2'
cost_inputs = {}
# Find rows with molecule names in column C
ws = wb_v[cost_sheet]
for r in range(1, ws.max_row + 1):
    tech = ws.cell(row=r, column=3).value  # col C
    if tech and isinstance(tech, str):
        tech_clean = tech.strip()
        if any(tech_clean == m['name'] for m in molecules):
            n_injections = cell_v(cost_sheet, f'E{r}')
            price = cell_v(cost_sheet, f'G{r}')
            cost_per_injection = cell_v(cost_sheet, f'I{r}')
            admin_cost = cell_v(cost_sheet, f'K{r}')
            cost_admin = cell_v(cost_sheet, f'M{r}')
            cost_total = cell_v(cost_sheet, f'O{r}')
            cost_annual_final = cell_v(cost_sheet, f'Q{r}')
            cost_inputs[tech_clean] = {
                'technology': tech_clean,
                'injectionCount': int(n_injections) if n_injections is not None else 0,
                'pricePerVial': float(price) if price else 0.0,
                'costPerInjection': float(cost_per_injection) if cost_per_injection else 0.0,
                'administrationCostPerVisit': float(admin_cost) if admin_cost else 0.0,
                'annualAdministrationCost': float(cost_admin) if cost_admin else 0.0,
                'annualTotalCost': float(cost_total) if cost_total else 0.0,
                'annualFinalCost': float(cost_annual_final) if cost_annual_final else 0.0,
            }

# ============================================================
# 6. Weights (from MCDA Advanced Dashboard)
# ============================================================
weights = []
mcda_sheet = 'MCDA Advanced Dashboard'
weight_rows = [
    (16, 'Cambio medio en la agudeza visual corregida (BCVA)', 'Eficacia'),
    (17, 'Pacientes con ganancia ≥15 letras en BCVA', 'Eficacia'),
    (19, 'Desprendimiento de retina', 'Seguridad'),
    (20, 'Infarto agudo de miocardio', 'Seguridad'),
    (21, 'Endoftalmitis', 'Seguridad'),
    (22, 'Supresión sistémica de VEGF', 'Seguridad'),
    (24, 'Calidad de vida relacionada con la salud', 'PROs'),
    (25, 'Discapacidad / capacidad funcional', 'PROs'),
    (27, 'Estado de aprobación regulatoria', 'Posicionamiento terapéutico'),
    (28, 'Inclusión en guías de práctica clínica', 'Posicionamiento terapéutico'),
]
for r, label, domain in weight_rows:
    w = cell_v(mcda_sheet, f'C{r}')
    weights.append({
        'domain': domain,
        'criteria': label,
        'weight': float(w) if w else 0.0,
    })

# ============================================================
# 7. References (from Referencias sheet)
# ============================================================
ref_sheet = 'Referencias'
references_by_molecule = {}
ws = wb_v[ref_sheet]
current_molecule = None
for r in range(1, ws.max_row + 1):
    a = cell_v(ref_sheet, f'A{r}')
    b = cell_v(ref_sheet, f'B{r}')
    if b and isinstance(b, str) and any(m['name'] == b.strip() for m in molecules):
        current_molecule = b.strip()
        references_by_molecule[current_molecule] = []
        continue
    if a and b and current_molecule:
        references_by_molecule[current_molecule].append({
            'domain': str(a).strip(),
            'citation': str(b).strip(),
        })

# ============================================================
# 8. T_results
# ============================================================
t_sheet = 'T_results'
t_results = []
for r in range(6, 14):
    name = cell_v(t_sheet, f'D{r}')
    if name:
        t_results.append({
            'molecule': str(name).strip(),
            'efficacy': float(cell_v(t_sheet, f'E{r}') or 0),
            'safety': float(cell_v(t_sheet, f'F{r}') or 0),
            'pros': float(cell_v(t_sheet, f'G{r}') or 0),
            'positioning': float(cell_v(t_sheet, f'H{r}') or 0),
            'totalGlobal': float(cell_v(t_sheet, f'I{r}') or 0),
            'annualCost': float(cell_v(t_sheet, f'J{r}') or 0),
            'costPerBenefit': float(cell_v(t_sheet, f'K{r}') or 0),
        })

# ============================================================
# 9. Build final seed data
# ============================================================
seed = {
    'meta': {
        'source': 'Modelo de valor degeneración macular asociada a la edad.xlsm',
        'description': 'Modelo de Valor - Análisis Multicriterio (MCDA) para degeneración macular asociada a la edad (AMD)',
        'version': '5.0',
        'discount': float(discount),
    },
    'molecules': molecules,
    'definitions': definitions,
    'parameters': parameters,
    'weights': weights,
    'moleculeInputs': molecule_inputs,
    'costInputs': cost_inputs,
    'references': references_by_molecule,
    'cachedResults': t_results,
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(seed, f, ensure_ascii=False, indent=2, default=str)

print(f'Saved seed data to {OUT}')
print(f'  Molecules: {len(molecules)}')
print(f'  Definitions: {len(definitions)}')
print(f'  Parameters: {len(parameters)}')
print(f'  Weights: {len(weights)}')
print(f'  Molecule inputs: {list(molecule_inputs.keys())}')
print(f'  Cost inputs: {len(cost_inputs)}')
print(f'  References: {sum(len(v) for v in references_by_molecule.values())} total')
print(f'  Cached results: {len(t_results)}')
print(f'  Discount: {discount}')

print('\n=== MOLECULES ===')
for m in molecules:
    print(f'  {m["index"]}: {m["name"]!r} ({m["columnRange"]})')

print('\n=== WEIGHTS ===')
for w in weights:
    print(f'  {w["domain"]:30s} | {w["criteria"][:50]:50s} | {w["weight"]}')

print('\n=== COST INPUTS ===')
for name, ci in cost_inputs.items():
    print(f'  {name!r}:')
    print(f'    inyecciones: {ci["injectionCount"]}, precio: {ci["pricePerVial"]:,}, total: {ci["annualTotalCost"]:,.0f}')
